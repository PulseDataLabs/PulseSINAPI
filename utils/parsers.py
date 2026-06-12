"""
utils/parsers.py
-----------------
Funções de parsing compartilhadas entre scrapers que consomem dados
de formatos heterogêneos (CSV, JSON, ZIP).
"""

import csv
import hashlib
import io
import json
import re
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
import warnings
import openpyxl
import xlrd

# Silencia avisos de formatação/estilos ausentes do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from .base import FUSO, agora_brt, get_logger, limpar

FIXOS = ["data_captura", "conjunto", "arquivo_origem", "registro_hash"]


def date_ref(type_date: str | None) -> datetime:
    now = datetime.now(FUSO)
    if type_date == "mes_anterior":
        return (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    return now


def replace_date_vars(url: str, dt: datetime) -> str:
    mapping = {
        "YYYY-MM": dt.strftime("%Y-%m"),
        "YYYY": dt.strftime("%Y"),
        "MM": dt.strftime("%m"),
    }
    for key, value in mapping.items():
        url = url.replace(key, value)
    return url


def decode_bytes(content: bytes) -> str:
    for enc in ("utf-8-sig", "latin1", "cp1252"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_key(k: str) -> str:
    key = limpar(k).lower().replace(" ", "_")
    key = unicodedata.normalize("NFKD", key)
    key = key.encode("ascii", "ignore").decode("ascii")
    key = re.sub(r"[^a-z0-9_]+", "_", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key or "campo"


def normalize_keys(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        out[normalize_key(str(k))] = limpar(v)
    return out


def csv_rows(text: str, delimiter: str | None = None) -> list[dict]:
    text = text.replace("\x00", "").strip()
    if not text:
        return []

    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []

    if delimiter is None:
        if ";" in lines[0] and "," not in lines[0]:
            delimiter = ";"
        else:
            delimiter = ","

    try:
        reader = csv.DictReader(io.StringIO("\n".join(lines)), delimiter=delimiter)
        rows = [normalize_keys(r) for r in reader if any((v or "").strip() for v in r.values())]
        if rows:
            return rows
    except Exception:
        pass

    return [{"linha": ln} for ln in lines]


def json_rows(payload) -> list[dict]:
    if isinstance(payload, dict):
        if isinstance(payload.get("results"), list):
            items = payload["results"]
        elif isinstance(payload.get("data"), list):
            items = payload["data"]
        else:
            items = [payload]
    elif isinstance(payload, list):
        items = payload
    else:
        items = [{"valor": str(payload)}]

    rows = []
    for it in items:
        if isinstance(it, dict):
            flat = {}
            for k, v in it.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = limpar(v)
            rows.append(normalize_keys(flat))
        else:
            rows.append({"valor": limpar(it)})
    return rows


def xls_rows(content: bytes) -> list[dict]:
    # openpyxl reader
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = [normalize_key(str(c.value or "")) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        has_value = False
        for c, key in enumerate(headers):
            if c >= len(row):
                break
            value = row[c]
            if isinstance(value, float):
                if value.is_integer():
                    value = str(int(value))
                else:
                    value = str(value)
            else:
                value = limpar(value)
            item[key] = value
            if value:
                has_value = True
        if has_value:
            rows.append(item)
    wb.close()
    return rows


def rows_from_zip(content: bytes) -> list[dict]:
    rows = []
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for info in zf.infolist():
            if info.is_dir() or info.file_size == 0:
                continue
            name = info.filename
            raw = zf.read(name)
            if name.lower().endswith((".csv", ".txt", ".tsv")):
                for row in csv_rows(decode_bytes(raw)):
                    row["arquivo_origem"] = name
                    rows.append(row)
            else:
                rows.append({"arquivo_origem": name, "tamanho_bytes": str(info.file_size)})
    return rows


def hash_row(row: dict) -> str:
    payload = json.dumps(row, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def read_existing_header(arquivo) -> list[str]:
    if not arquivo.exists():
        return []
    with arquivo.open(newline="", encoding="utf-8") as f:
        first = f.readline().strip()
    return [c.strip() for c in first.split(",") if c.strip()] if first else []


def enriquecer(dataset_id: str, rows: list[dict]) -> tuple[list[dict], list[str]]:
    data_captura, _ = agora_brt()
    enriched = []
    campos = set()
    for row in rows:
        item = {k: limpar(v) for k, v in row.items()}
        item["data_captura"] = data_captura
        item["conjunto"] = dataset_id
        item.setdefault("arquivo_origem", "")
        item["registro_hash"] = hash_row({k: v for k, v in item.items() if k != "data_captura"})
        campos.update(item.keys())
        enriched.append(item)

    ordenados = [c for c in sorted(campos) if c not in FIXOS]
    header = FIXOS + ordenados
    return enriched, header
